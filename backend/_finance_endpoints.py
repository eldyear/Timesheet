
# ============================================================
# FINANCE MODULE — appended by migrate script
# ============================================================

def _require_finance_view(current_user: models.User = Depends(get_current_user)):
    if not getattr(current_user.role, 'can_view_finance', False) and not current_user.role.can_manage_settings:
        raise HTTPException(status_code=403, detail="Finance access required")
    return current_user

def _require_finance_edit(current_user: models.User = Depends(get_current_user)):
    if not getattr(current_user.role, 'can_edit_finance', False) and not current_user.role.can_manage_settings:
        raise HTTPException(status_code=403, detail="Finance edit permission required")
    return current_user

def _write_audit(db: Session, user: models.User, action: str, target: str,
                 old_value: str = None, new_value: str = None):
    log = models.FinanceAuditLog(
        user_id=user.id, action=action, target=target,
        old_value=old_value, new_value=new_value,
        timestamp=datetime.utcnow()
    )
    db.add(log)


@app.get("/api/salary-rates")
def get_salary_rates(db: Session = Depends(get_db),
                     current_user: models.User = Depends(_require_finance_view)):
    rates = db.query(models.SalaryRate).all()
    return [{"id": r.id, "dept_id": r.dept_id,
             "dept_name": r.department.name if r.department else None,
             "position_id": r.position_id,
             "position_name": r.position.name if r.position else None,
             "hourly_rate": r.hourly_rate} for r in rates]


@app.post("/api/salary-rates")
def upsert_salary_rate(rate: SalaryRateCreate, db: Session = Depends(get_db),
                       current_user: models.User = Depends(_require_finance_edit)):
    existing = db.query(models.SalaryRate).filter(
        models.SalaryRate.dept_id == rate.dept_id,
        models.SalaryRate.position_id == rate.position_id
    ).first()
    pos = db.query(models.Position).get(rate.position_id)
    dept = db.query(models.Department).get(rate.dept_id)
    target = f"{pos.name if pos else rate.position_id} in {dept.name if dept else rate.dept_id}"
    if existing:
        old_val = str(existing.hourly_rate)
        existing.hourly_rate = rate.hourly_rate
        _write_audit(db, current_user, "UPDATE_RATE", target, old_val, str(rate.hourly_rate))
        db.commit()
        db.refresh(existing)
        r = existing
    else:
        r = models.SalaryRate(**rate.dict())
        db.add(r)
        _write_audit(db, current_user, "CREATE_RATE", target, None, str(rate.hourly_rate))
        db.commit()
        db.refresh(r)
    return {"id": r.id, "dept_id": r.dept_id, "position_id": r.position_id, "hourly_rate": r.hourly_rate}


@app.delete("/api/salary-rates/{rate_id}")
def delete_salary_rate(rate_id: int, db: Session = Depends(get_db),
                       current_user: models.User = Depends(_require_finance_edit)):
    r = db.query(models.SalaryRate).filter(models.SalaryRate.id == rate_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Rate not found")
    target = f"position_id={r.position_id} dept_id={r.dept_id}"
    _write_audit(db, current_user, "DELETE_RATE", target, str(r.hourly_rate), None)
    db.delete(r)
    db.commit()
    return {"status": "deleted"}


@app.get("/api/finance/payroll/{year_month}")
def get_payroll(year_month: str, db: Session = Depends(get_db),
                current_user: models.User = Depends(_require_finance_view)):
    try:
        year, month = map(int, year_month.split("-"))
    except ValueError:
        raise HTTPException(status_code=400, detail="year_month must be YYYY-MM")
    _, days = calendar.monthrange(year, month)
    month_start = date(year, month, 1)
    month_end = date(year, month, days)
    rates = db.query(models.SalaryRate).all()
    rate_map = {(r.dept_id, r.position_id): r.hourly_rate for r in rates}
    employees = db.query(models.Employee).all()
    rows = []
    dept_totals: Dict[int, dict] = {}
    for emp in employees:
        entries = db.query(models.Timesheet).filter(
            models.Timesheet.employee_id == emp.id,
            models.Timesheet.date >= month_start,
            models.Timesheet.date <= month_end,
        ).all()
        rate = rate_map.get((emp.dept_id, emp.position_id), 0.0)
        total_std = 0.0
        total_night = 0.0
        gross_pay = 0.0
        for entry in entries:
            wc = db.query(models.WorkCode).get(entry.work_code_id)
            if wc:
                m = wc.rate_multiplier or 1.0
                total_std += wc.hours_standard
                total_night += wc.hours_night
                gross_pay += (wc.hours_standard + wc.hours_night) * rate * m
        dept_name = emp.department.name if emp.department else "Unknown"
        pos_name = emp.position.name if emp.position else "\u2014"
        rows.append({
            "employee_id": emp.id, "full_name": emp.full_name,
            "tab_number": emp.tab_number, "position": pos_name,
            "dept_id": emp.dept_id, "dept_name": dept_name,
            "hourly_rate": rate,
            "std_hours": round(total_std, 1), "night_hours": round(total_night, 1),
            "total_hours": round(total_std + total_night, 1),
            "gross_pay": round(gross_pay, 2),
        })
        if emp.dept_id not in dept_totals:
            dept_totals[emp.dept_id] = {"dept_name": dept_name, "total_pay": 0.0, "employees": 0}
        dept_totals[emp.dept_id]["total_pay"] += gross_pay
        dept_totals[emp.dept_id]["employees"] += 1
    grand_total = sum(r["gross_pay"] for r in rows)
    avg_salary = round(grand_total / len(rows), 2) if rows else 0.0
    dept_summary = [{"dept_id": did, **info, "total_pay": round(info["total_pay"], 2)}
                    for did, info in dept_totals.items()]
    top = max(dept_summary, key=lambda x: x["total_pay"]) if dept_summary else None
    return {
        "year_month": year_month, "employees": rows, "dept_summary": dept_summary,
        "grand_total": round(grand_total, 2), "avg_salary": avg_salary,
        "top_dept": top["dept_name"] if top else None,
        "top_dept_pay": round(top["total_pay"], 2) if top else 0.0,
    }


@app.get("/api/finance/payroll/{year_month}/export")
def export_payroll_excel(year_month: str, db: Session = Depends(get_db),
                         current_user: models.User = Depends(_require_finance_view)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    payroll = get_payroll(year_month, db, current_user)
    wb = Workbook()
    ws = wb.active
    ws.title = f"Payroll {year_month}"
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(fill_type="solid", fgColor="4F46E5")
    alt = PatternFill(fill_type="solid", fgColor="F1F5F9")
    hdrs = ["#", "Tab No.", "Full Name", "Position", "Department",
            "Rate/hr", "Std Hrs", "Night Hrs", "Total Hrs", "Gross Pay"]
    ws.append(hdrs)
    for c in range(1, len(hdrs) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hf; cell.fill = hfill
        cell.alignment = Alignment(horizontal="center"); cell.border = border
    for i, emp in enumerate(payroll["employees"], 1):
        rdata = [i, emp["tab_number"], emp["full_name"], emp["position"], emp["dept_name"],
                 emp["hourly_rate"], emp["std_hours"], emp["night_hours"],
                 emp["total_hours"], emp["gross_pay"]]
        ws.append(rdata)
        fill = alt if i % 2 == 0 else None
        for c in range(1, len(rdata) + 1):
            cell = ws.cell(row=i + 1, column=c)
            cell.border = border
            if fill: cell.fill = fill
            if c == 10: cell.number_format = '#,##0.00'
    tr = ws.max_row + 1
    ws.cell(tr, 1, "TOTAL").font = Font(bold=True)
    tc = ws.cell(tr, 10, payroll["grand_total"])
    tc.font = Font(bold=True); tc.number_format = '#,##0.00'
    tc.fill = PatternFill(fill_type="solid", fgColor="C7D2FE")
    for c, w in enumerate([4, 10, 28, 20, 20, 10, 9, 10, 10, 14], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=payroll_{year_month}.xlsx"})


@app.get("/api/finance/audit-log")
def get_audit_log(db: Session = Depends(get_db),
                  current_user: models.User = Depends(_require_finance_edit)):
    logs = db.query(models.FinanceAuditLog).order_by(
        models.FinanceAuditLog.timestamp.desc()).limit(200).all()
    return [{
        "id": log.id,
        "username": log.user.username if log.user else "system",
        "action": log.action, "target": log.target,
        "old_value": log.old_value, "new_value": log.new_value,
        "timestamp": log.timestamp.isoformat() if log.timestamp else None,
    } for log in logs]
