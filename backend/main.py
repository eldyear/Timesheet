from fastapi import FastAPI, Depends, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import extract
import models
from database import SessionLocal, engine
from pydantic import BaseModel
from typing import List, Optional, Dict
from datetime import date
import calendar
from datetime import datetime, timedelta
import io
import pandas as pd
from fastapi.responses import StreamingResponse
from openpyxl.utils import get_column_letter

from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from passlib.context import CryptContext
from jose import JWTError, jwt

# --- Security Config ---
SECRET_KEY = "super-secret-key-for-mvp" # In production, use environment variable
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7 # 1 week

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login")

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

# Создаем таблицы при запуске
models.Base.metadata.create_all(bind=engine)

# Inline migration: add new role permission columns if they don't exist yet
def _run_startup_migrations():
    new_role_cols = [
        ("can_export", "BOOLEAN DEFAULT FALSE"),
        ("can_manage_employees", "BOOLEAN DEFAULT FALSE"),
        ("can_manage_users", "BOOLEAN DEFAULT FALSE"),
        ("can_manage_departments", "BOOLEAN DEFAULT FALSE"),
    ]
    with engine.connect() as conn:
        for col_name, col_def in new_role_cols:
            try:
                conn.execute(text(f"ALTER TABLE roles ADD COLUMN IF NOT EXISTS {col_name} {col_def}"))
                conn.commit()
            except Exception:
                conn.rollback()

_run_startup_migrations()

app = FastAPI(title="Timesheet API")

# Add CORS for React frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], # Allow all for MVP
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- Pydantic Schemas ---
class DepartmentCreate(BaseModel):
    name: str
    parent_id: Optional[int] = None
    category: int = 99

class DepartmentSchema(DepartmentCreate):
    id: int
    
    class Config:
        from_attributes = True

class WorkCodeCreate(BaseModel):
    code: str
    label: str
    hours_standard: float = 0.0
    hours_night: float = 0.0
    color_hex: str = "#FFFFFF"
    rate_multiplier: float = 1.0

class WorkCodeSchema(WorkCodeCreate):
    id: int
    
    class Config:
        from_attributes = True

class PositionCreate(BaseModel):
    name: str

class PositionSchema(PositionCreate):
    id: int
    class Config:
        from_attributes = True

class EmployeeCreate(BaseModel):
    full_name: str
    tab_number: str
    category: int = 99
    position_id: Optional[int] = None
    dept_id: int

class EmployeeSchema(BaseModel):
    id: int
    full_name: str
    tab_number: str
    category: int
    tab_number: str
    position_id: Optional[int] = None
    position: Optional[PositionSchema] = None
    dept_id: int
    
    class Config:
        from_attributes = True

class RoleCreate(BaseModel):
    name: str
    can_manage_settings: bool = False
    can_edit_all: bool = False
    can_view_all: bool = False
    can_view_only: bool = True
    can_view_finance: bool = False
    can_edit_finance: bool = False
    can_export: bool = False
    can_manage_employees: bool = False
    can_manage_users: bool = False
    can_manage_departments: bool = False

class RoleUpdate(BaseModel):
    name: Optional[str] = None
    can_manage_settings: Optional[bool] = None
    can_edit_all: Optional[bool] = None
    can_view_all: Optional[bool] = None
    can_view_only: Optional[bool] = None
    can_view_finance: Optional[bool] = None
    can_edit_finance: Optional[bool] = None
    can_export: Optional[bool] = None
    can_manage_employees: Optional[bool] = None
    can_manage_users: Optional[bool] = None
    can_manage_departments: Optional[bool] = None

class RoleSchema(RoleCreate):
    id: int
    
    class Config:
        from_attributes = True

class SalaryRateCreate(BaseModel):
    dept_id: int
    position_id: int
    hourly_rate: float

class SalaryRateSchema(SalaryRateCreate):
    id: int
    dept_name: Optional[str] = None
    position_name: Optional[str] = None

    class Config:
        from_attributes = True

class AuditLogSchema(BaseModel):
    id: int
    username: Optional[str] = None
    action: str
    target: str
    old_value: Optional[str] = None
    new_value: Optional[str] = None
    timestamp: str

    class Config:
        from_attributes = True

class UserCreate(BaseModel):
    username: str
    password: str
    role_id: int
    dept_id: Optional[int] = None
    employee_id: Optional[int] = None

class UserUpdate(BaseModel):
    username: Optional[str] = None
    password: Optional[str] = None
    role_id: Optional[int] = None
    dept_id: Optional[int] = None
    employee_id: Optional[int] = None

class UserSchema(BaseModel):
    id: int
    username: str
    role_id: int
    role: Optional[RoleSchema] = None
    dept_id: Optional[int] = None
    employee_id: Optional[int] = None
    
    class Config:
        from_attributes = True

class Token(BaseModel):
    access_token: str
    token_type: str

class TokenData(BaseModel):
    username: Optional[str] = None
    role: Optional[dict] = None
    dept_id: Optional[int] = None

class TimesheetUpdateItem(BaseModel):
    employee_id: int
    date: date
    work_code_id: Optional[int]

class TimesheetUpdateRequest(BaseModel):
    updates: List[TimesheetUpdateItem]

# Получение сессии БД
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=401,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
        token_data = TokenData(username=username, role=payload.get("role"), dept_id=payload.get("dept_id"))
    except JWTError:
        raise credentials_exception
    user = db.query(models.User).filter(models.User.username == token_data.username).first()
    if user is None:
        raise credentials_exception
    return user

@app.post("/api/auth/login", response_model=Token)
def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.username == form_data.username).first()
    if not user or not verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=401,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    
    role_dict = {
        "id": user.role.id,
        "name": user.role.name,
        "can_manage_settings": user.role.can_manage_settings,
        "can_edit_all": user.role.can_edit_all,
        "can_view_all": user.role.can_view_all,
        "can_view_only": user.role.can_view_only,
        "can_view_finance": user.role.can_view_finance,
        "can_edit_finance": user.role.can_edit_finance,
        "can_export": user.role.can_export if hasattr(user.role, 'can_export') else False,
        "can_manage_employees": user.role.can_manage_employees if hasattr(user.role, 'can_manage_employees') else False,
        "can_manage_users": user.role.can_manage_users if hasattr(user.role, 'can_manage_users') else False,
        "can_manage_departments": user.role.can_manage_departments if hasattr(user.role, 'can_manage_departments') else False,
    }
    
    active_dept_id = user.employee.dept_id if user.employee else user.dept_id
    
    access_token = create_access_token(
        data={"sub": user.username, "role": role_dict, "dept_id": active_dept_id},
        expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}

@app.get("/")
def read_root():
    return {"message": "Timesheet API is running"}

@app.get("/api/users", response_model=List[UserSchema])
def get_users(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if not current_user.role.can_manage_settings:
        raise HTTPException(status_code=403, detail="Only admins can perform this action")
    return db.query(models.User).all()

@app.post("/api/users", response_model=UserSchema)
def create_user(user: UserCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if not current_user.role.can_manage_settings:
        raise HTTPException(status_code=403, detail="Only admins can perform this action")
    
    hashed_password = get_password_hash(user.password)
    db_user = models.User(username=user.username, hashed_password=hashed_password, role_id=user.role_id, dept_id=user.dept_id, employee_id=user.employee_id)
    try:
        db.add(db_user)
        db.commit()
        db.refresh(db_user)
        return db_user
    except Exception:
        db.rollback()
        raise HTTPException(status_code=400, detail="Username already exists")

@app.put("/api/users/{user_id}", response_model=UserSchema)
def update_user(user_id: int, user: UserUpdate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if not current_user.role.can_manage_settings:
        raise HTTPException(status_code=403, detail="Only admins can perform this action")
    
    db_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")
        
    for key, value in user.dict(exclude_unset=True).items():
        if key == "password" and value:
            db_user.hashed_password = get_password_hash(value)
        elif key != "password":
            setattr(db_user, key, value)
            
    db.commit()
    db.refresh(db_user)
    return db_user

@app.delete("/api/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if not current_user.role.can_manage_settings:
        raise HTTPException(status_code=403, detail="Only admins can perform this action")
        
    db_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")
        
    if db_user.username == "admin":
        raise HTTPException(status_code=400, detail="Cannot delete default admin user")
        
    db.delete(db_user)
    db.commit()
    return {"status": "deleted"}

@app.get("/api/roles", response_model=List[RoleSchema])
def get_roles(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if not current_user.role.can_manage_settings:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    return db.query(models.Role).all()

@app.post("/api/roles", response_model=RoleSchema)
def create_role(role: RoleCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if not current_user.role.can_manage_settings:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    db_role = models.Role(**role.dict())
    db.add(db_role)
    db.commit()
    db.refresh(db_role)
    return db_role

@app.put("/api/roles/{role_id}", response_model=RoleSchema)
def update_role(role_id: int, role: RoleUpdate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if not current_user.role.can_manage_settings:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    db_role = db.query(models.Role).filter(models.Role.id == role_id).first()
    if not db_role:
        raise HTTPException(status_code=404, detail="Role not found")
    for key, value in role.dict(exclude_unset=True).items():
        setattr(db_role, key, value)
    db.commit()
    db.refresh(db_role)
    return db_role

@app.delete("/api/roles/{role_id}")
def delete_role(role_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if not current_user.role.can_manage_settings:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    db_role = db.query(models.Role).filter(models.Role.id == role_id).first()
    if not db_role:
        raise HTTPException(status_code=404, detail="Role not found")
    
    # Prevent deletion if users are assigned to it
    users_with_role = db.query(models.User).filter(models.User.role_id == role_id).first()
    if users_with_role:
        raise HTTPException(status_code=400, detail="Cannot delete role while users are assigned to it")
        
    db.delete(db_role)
    db.commit()
    return {"status": "deleted"}

def _get_root_department_id(db: Session, dept_id: Optional[int]) -> Optional[int]:
    if dept_id is None:
        return None
    current = db.query(models.Department).filter(models.Department.id == dept_id).first()
    if not current:
        return dept_id
    while current and current.parent_id is not None:
        parent = db.query(models.Department).filter(models.Department.id == current.parent_id).first()
        if parent is None:
            break
        current = parent
    return current.id

@app.get("/api/departments", response_model=List[DepartmentSchema])
def get_departments(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    """Returns departments based on access rights."""
    if current_user.role.can_view_all or current_user.role.can_edit_all or current_user.role.can_manage_settings:
        return db.query(models.Department).all()
    
    active_dept_id = current_user.active_dept_id
    if active_dept_id is None:
        # User with no department assignment — return all visible (unscoped)
        return db.query(models.Department).all()
        
    root_id = _get_root_department_id(db, active_dept_id)
    allowed_ids = _get_department_hierarchy_ids(db, root_id)
    return db.query(models.Department).filter(models.Department.id.in_(allowed_ids)).order_by(models.Department.category, models.Department.id).all()

@app.post("/api/departments", response_model=DepartmentSchema)
def create_department(dept: DepartmentCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    can = current_user.role.can_manage_settings or (current_user.role.can_manage_departments if hasattr(current_user.role, 'can_manage_departments') else False)
    if not can:
        raise HTTPException(status_code=403, detail="Not authorized to manage departments")
    db_dept = models.Department(**dept.dict())
    db.add(db_dept)
    db.commit()
    db.refresh(db_dept)
    return db_dept

@app.put("/api/departments/{dept_id}", response_model=DepartmentSchema)
def update_department(dept_id: int, dept: DepartmentCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    can = current_user.role.can_manage_settings or (current_user.role.can_manage_departments if hasattr(current_user.role, 'can_manage_departments') else False)
    if not can:
        raise HTTPException(status_code=403, detail="Not authorized to manage departments")
    db_dept = db.query(models.Department).filter(models.Department.id == dept_id).first()
    if not db_dept:
        raise HTTPException(status_code=404, detail="Department not found")
    for key, value in dept.dict().items():
        setattr(db_dept, key, value)
    db.commit()
    db.refresh(db_dept)
    return db_dept

@app.delete("/api/departments/{dept_id}")
def delete_department(dept_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if not current_user.role.can_manage_settings:
        raise HTTPException(status_code=403, detail="Only admins can perform this action")
    db_dept = db.query(models.Department).filter(models.Department.id == dept_id).first()
    if not db_dept:
        raise HTTPException(status_code=404, detail="Department not found")
        
    # Check for child departments
    children = db.query(models.Department).filter(models.Department.parent_id == dept_id).first()
    if children:
        raise HTTPException(status_code=400, detail="Cannot delete department with sub-departments")
        
    # Check for assigned employees
    employees = db.query(models.Employee).filter(models.Employee.dept_id == dept_id).first()
    if employees:
        raise HTTPException(status_code=400, detail="Cannot delete department with assigned employees")
        
    db.delete(db_dept)
    db.commit()
    return {"status": "deleted"}

# --- Positions CRUD ---
@app.get("/api/positions", response_model=List[PositionSchema])
def get_positions(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    return db.query(models.Position).order_by(models.Position.name).all()

@app.post("/api/positions", response_model=PositionSchema)
def create_position(pos: PositionCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if not current_user.role.can_manage_settings:
        raise HTTPException(status_code=403, detail="Only admins can perform this action")
    db_pos = models.Position(**pos.dict())
    try:
        db.add(db_pos)
        db.commit()
        db.refresh(db_pos)
        return db_pos
    except Exception:
        db.rollback()
        raise HTTPException(status_code=400, detail="Position name must be unique")

@app.put("/api/positions/{pos_id}", response_model=PositionSchema)
def update_position(pos_id: int, pos: PositionCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if not current_user.role.can_manage_settings:
        raise HTTPException(status_code=403, detail="Only admins can perform this action")
    db_pos = db.query(models.Position).filter(models.Position.id == pos_id).first()
    if not db_pos:
        raise HTTPException(status_code=404, detail="Position not found")
    db_pos.name = pos.name
    db.commit()
    db.refresh(db_pos)
    return db_pos

@app.delete("/api/positions/{pos_id}")
def delete_position(pos_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if not current_user.role.can_manage_settings:
        raise HTTPException(status_code=403, detail="Only admins can perform this action")
    db_pos = db.query(models.Position).filter(models.Position.id == pos_id).first()
    if not db_pos:
        raise HTTPException(status_code=404, detail="Position not found")
    assigned = db.query(models.Employee).filter(models.Employee.position_id == pos_id).first()
    if assigned:
        raise HTTPException(status_code=400, detail="Cannot delete position with assigned employees")
    db.delete(db_pos)
    db.commit()
    return {"status": "deleted"}

@app.get("/api/work-codes", response_model=List[WorkCodeSchema])
def get_work_codes(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    """Returns the list of available work codes (marks)."""
    return db.query(models.WorkCode).all()

@app.post("/api/work-codes", response_model=WorkCodeSchema)
def create_work_code(wc: WorkCodeCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if not current_user.role.can_manage_settings:
        raise HTTPException(status_code=403, detail="Only admins can perform this action")
    db_wc = models.WorkCode(**wc.dict())
    try:
        db.add(db_wc)
        db.commit()
        db.refresh(db_wc)
        return db_wc
    except Exception:
        db.rollback()
        raise HTTPException(status_code=400, detail="Work code already exists")

@app.put("/api/work-codes/{wc_id}", response_model=WorkCodeSchema)
def update_work_code(wc_id: int, wc: WorkCodeCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if not current_user.role.can_manage_settings:
        raise HTTPException(status_code=403, detail="Only admins can perform this action")
    db_wc = db.query(models.WorkCode).filter(models.WorkCode.id == wc_id).first()
    if not db_wc:
        raise HTTPException(status_code=404, detail="Work code not found")
    for key, value in wc.dict().items():
        setattr(db_wc, key, value)
    try:
        db.commit()
        db.refresh(db_wc)
        return db_wc
    except Exception:
        db.rollback()
        raise HTTPException(status_code=400, detail="Database Error updating Work Code")

@app.delete("/api/work-codes/{wc_id}")
def delete_work_code(wc_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if not current_user.role.can_manage_settings:
        raise HTTPException(status_code=403, detail="Only admins can perform this action")
    db_wc = db.query(models.WorkCode).filter(models.WorkCode.id == wc_id).first()
    if not db_wc:
        raise HTTPException(status_code=404, detail="Work code not found")
    db.delete(db_wc)
    db.commit()
    return {"status": "deleted"}

@app.get("/api/employees/next-tab-number")
def get_next_tab_number(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    """Returns the next available tab number as a zero-padded string (max existing + 1)."""
    from sqlalchemy import func
    employees = db.query(models.Employee.tab_number).all()
    max_num = 0
    for (tn,) in employees:
        try:
            n = int(tn)
            if n > max_num:
                max_num = n
        except (ValueError, TypeError):
            pass
    next_num = max_num + 1
    return {"next_tab_number": str(next_num).zfill(3)}

@app.get("/api/employees", response_model=List[EmployeeSchema])

def get_employees(dept_id: Optional[int] = None, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    is_global = current_user.role.can_view_all or current_user.role.can_edit_all or current_user.role.can_manage_settings
    if is_global:
        if dept_id:
            allowed_ids = _get_department_hierarchy_ids(db, dept_id)
            return db.query(models.Employee).filter(models.Employee.dept_id.in_(allowed_ids)).all()
        return db.query(models.Employee).all()
    else:
        active_dept_id = current_user.active_dept_id
        if active_dept_id is None:
            return db.query(models.Employee).all()
        root_id = _get_root_department_id(db, active_dept_id)
        if dept_id:
            allowed_root_ids = _get_department_hierarchy_ids(db, root_id)
            if dept_id not in allowed_root_ids:
                return []
            allowed_ids = _get_department_hierarchy_ids(db, dept_id)
            return db.query(models.Employee).filter(models.Employee.dept_id.in_(allowed_ids)).all()
        else:
            allowed_ids = _get_department_hierarchy_ids(db, root_id)
            return db.query(models.Employee).filter(models.Employee.dept_id.in_(allowed_ids)).all()

@app.post("/api/employees", response_model=EmployeeSchema)
def create_employee(emp: EmployeeCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    can = current_user.role.can_manage_settings or (current_user.role.can_manage_employees if hasattr(current_user.role, 'can_manage_employees') else False)
    if not can:
        raise HTTPException(status_code=403, detail="Not authorized to manage employees")
    db_emp = models.Employee(**emp.dict())
    try:
        db.add(db_emp)
        db.commit()
        db.refresh(db_emp)
        return db_emp
    except Exception:
        db.rollback()
        raise HTTPException(status_code=400, detail="Employee with that Tab Number might exist")

@app.put("/api/employees/{emp_id}", response_model=EmployeeSchema)
def update_employee(emp_id: int, emp: EmployeeCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    can = current_user.role.can_manage_settings or (current_user.role.can_manage_employees if hasattr(current_user.role, 'can_manage_employees') else False)
    if not can:
        raise HTTPException(status_code=403, detail="Not authorized to manage employees")
    db_emp = db.query(models.Employee).filter(models.Employee.id == emp_id).first()
    if not db_emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    for key, value in emp.dict().items():
        setattr(db_emp, key, value)
    db.commit()
    db.refresh(db_emp)
    return db_emp

@app.delete("/api/employees/{emp_id}")
def delete_employee(emp_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if not current_user.role.can_manage_settings:
        raise HTTPException(status_code=403, detail="Only admins can perform this action")
    db_emp = db.query(models.Employee).filter(models.Employee.id == emp_id).first()
    if not db_emp:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # Check for timesheet entries to prevent foreign key constraint violations
    timesheets = db.query(models.Timesheet).filter(models.Timesheet.employee_id == emp_id).first()
    if timesheets:
        raise HTTPException(status_code=400, detail="Cannot delete employee with existing timesheet records")
        
    db.delete(db_emp)
    db.commit()
    return {"status": "deleted"}

def _get_department_hierarchy_ids(db: Session, dept_id: int) -> List[int]:
    ids = [dept_id]
    children = db.query(models.Department).filter(models.Department.parent_id == dept_id).all()
    for child in children:
        ids.extend(_get_department_hierarchy_ids(db, child.id))
    return ids

@app.get("/api/timesheet/{dept_id}/{year_month}")
def get_timesheet(dept_id: int, year_month: str, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    """
    Returns a grid-ready JSON containing employees and their existing marks for the specified month.
    year_month format: YYYY-MM
    """
    if not current_user.role.can_view_all and not current_user.role.can_edit_all:
        allowed_dept_ids = _get_department_hierarchy_ids(db, current_user.active_dept_id)
        if dept_id not in allowed_dept_ids:
            raise HTTPException(status_code=403, detail="Not authorized to view this department's timesheet")

    try:
        year, month = map(int, year_month.split("-"))
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid month format. Expected YYYY-MM")

    # 1. Get employees in the department and sub-departments
    dept_ids = _get_department_hierarchy_ids(db, dept_id)
    employees = db.query(models.Employee).filter(models.Employee.dept_id.in_(dept_ids)).all()
    emp_dict = [EmployeeSchema.from_orm(emp).dict() for emp in employees]
    emp_ids = [emp.id for emp in employees]

    # 2. Get timesheet entries for these employees for this month
    start_date = date(year, month, 1)
    _, last_day = calendar.monthrange(year, month)
    end_date = date(year, month, last_day)

    entries = db.query(models.Timesheet).filter(
        models.Timesheet.employee_id.in_(emp_ids),
        models.Timesheet.date >= start_date,
        models.Timesheet.date <= end_date
    ).all()

    # 3. Format as nested dictionary: employee_id -> day -> work_code_id
    timesheet_data = {emp.id: {} for emp in employees}
    for entry in entries:
        day = entry.date.day
        timesheet_data[entry.employee_id][day] = entry.work_code_id

    return {
        "employees": emp_dict,
        "timesheet": timesheet_data,
        "days_in_month": last_day,
        "month": month,
        "year": year
    }

@app.get("/api/export/t13/{dept_id}/{year_month}")
def export_t13(dept_id: int, year_month: str, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    """Exports Timesheet to Excel T-13 Format with Department Grouping"""
    if not current_user.role.can_view_all and not current_user.role.can_edit_all and not current_user.role.can_view_only:
        allowed_dept_ids = _get_department_hierarchy_ids(db, current_user.active_dept_id)
        if dept_id not in allowed_dept_ids:
            raise HTTPException(status_code=403, detail="Not authorized to export this department's timesheet")

    try:
        year, month = map(int, year_month.split("-"))
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid month format. Expected YYYY-MM")

    department = db.query(models.Department).filter(models.Department.id == dept_id).first()
    if not department:
        raise HTTPException(status_code=404, detail="Department not found")

    dept_ids = _get_department_hierarchy_ids(db, dept_id)
    departments = db.query(models.Department).filter(models.Department.id.in_(dept_ids)).all()
    employees = db.query(models.Employee).filter(models.Employee.dept_id.in_(dept_ids)).all()
    emp_ids = [emp.id for emp in employees]

    _, last_day = calendar.monthrange(year, month)
    start_date = date(year, month, 1)
    end_date = date(year, month, last_day)

    entries = db.query(models.Timesheet).filter(
        models.Timesheet.employee_id.in_(emp_ids),
        models.Timesheet.date >= start_date,
        models.Timesheet.date <= end_date
    ).all()

    work_codes = {wc.id: wc for wc in db.query(models.WorkCode).all()}
    
    # Timesheet Map: emp.id -> day -> wc
    timesheet_map = {emp.id: {} for emp in employees}
    for entry in entries:
        timesheet_map[entry.employee_id][entry.date.day] = work_codes.get(entry.work_code_id)

    # Calculate employee totals
    totals = {}
    for emp in employees:
        std = 0.0
        night = 0.0
        for day in range(1, last_day + 1):
            wc = timesheet_map[emp.id].get(day)
            if wc:
                std += (wc.hours_standard or 0.0)
                night += (wc.hours_night or 0.0)
        totals[emp.id] = {"std": round(std, 1), "night": round(night, 1), "total": round(std + night, 1)}

    # Group employees by department for rendering
    depts_map = {d.id: d for d in departments}
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from urllib.parse import quote
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Timesheet {year_month}"
    
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Header Font and Fill
    hf = Font(bold=True, color="1E293B")
    hfill = PatternFill(fill_type="solid", fgColor="F8FAFC")
    
    # Headers
    hdrs = ["Employee Name", "Tab No.", "Position", "Category"]
    for day in range(1, last_day + 1):
        hdrs.append(str(day))
    hdrs.extend(["Std Hrs", "Night Hrs", "Total Hrs"])
    
    ws.append(hdrs)
    for c in range(1, len(hdrs) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Process Departments Top-Down
    # Sort departments by category, then by ID
    sorted_depts = sorted(departments, key=lambda d: (d.category or 99, d.id))
    
    current_row = 2
    
    for dept in sorted_depts:
        dept_emps = [e for e in employees if e.dept_id == dept.id]
        if not dept_emps:
            continue
            
        # Draw Department Banner
        dept_cell = ws.cell(row=current_row, column=1)
        
        parent_name = ""
        if dept.parent_id and dept.parent_id in depts_map:
            parent_name = f"{depts_map[dept.parent_id].name} » "
            
        dept_cell.value = f"{parent_name}{dept.name}"
        dept_cell.font = Font(bold=True, color="0F172A")
        dept_cell.fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
        
        # Merge department banner across all columns
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(hdrs))
        # Apply borders to merged cell
        for c in range(1, len(hdrs) + 1):
            ws.cell(row=current_row, column=c).border = border
            
        current_row += 1
        
        # Draw Employees
        dept_emps.sort(key=lambda e: (e.category or 99, e.full_name))
        for emp in dept_emps:
            pos_name = emp.position.name if emp.position else "—"
            cat_name = str(emp.category) if emp.category is not None else "99"
            row_data = [emp.full_name, emp.tab_number, pos_name, cat_name]
            
            for day in range(1, last_day + 1):
                wc = timesheet_map[emp.id].get(day)
                row_data.append(wc.code if wc else "")
                
            emp_totals = totals[emp.id]
            row_data.extend([emp_totals["std"], emp_totals["night"], emp_totals["total"]])
            
            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=c_idx)
                cell.value = val
                cell.border = border
                if c_idx > 4:  # Center align days and totals
                    cell.alignment = Alignment(horizontal="center")
                    
            current_row += 1

    # Adjust Column Widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    for day in range(1, last_day + 1):
        ws.column_dimensions[get_column_letter(day + 4)].width = 5
    for c in range(1, 4):
        ws.column_dimensions[get_column_letter(last_day + 4 + c)].width = 10

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"T-13_{department.full_name.replace(' ', '_').replace('»', '-')}_{year_month}.xlsx"
    encoded_filename = quote(filename)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=utf-8''{encoded_filename}"}
    )

@app.post("/api/timesheet/update")
def update_timesheet(payload: TimesheetUpdateRequest, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    """Bulk update endpoint to save changes from the grid."""
    
    # Optional Security: For the MVP, we assume the frontend sends Employee IDs within the user's allowed department.
    # In a full app, we would query the Employee IDs and verify all belong to current_user.dept_id.
    if not current_user.role.can_edit_all:
        allowed_dept_ids = _get_department_hierarchy_ids(db, current_user.active_dept_id)
        for item in payload.updates:
            emp = db.query(models.Employee).get(item.employee_id)
            if emp and emp.dept_id not in allowed_dept_ids:
                raise HTTPException(status_code=403, detail="Cannot edit employees outside your department")
    
    for item in payload.updates:
        # Check if entry exists
        existing = db.query(models.Timesheet).filter(
            models.Timesheet.employee_id == item.employee_id,
            models.Timesheet.date == item.date
        ).first()

        if item.work_code_id is None:
            # Delete entry if exists
            if existing:
                db.delete(existing)
        else:
            if existing:
                # Update
                existing.work_code_id = item.work_code_id
            else:
                # Insert
                new_entry = models.Timesheet(
                    employee_id=item.employee_id,
                    date=item.date,
                    work_code_id=item.work_code_id
                )
                db.add(new_entry)
                
    db.commit()
    return {"status": "success", "updated_count": len(payload.updates)}
# ============================================================
# FINANCE MODULE — appended by migrate script
# ============================================================

def _require_finance_view(current_user: models.User = Depends(get_current_user)):
    if not getattr(current_user.role, 'can_view_finance', False) and not current_user.role.can_manage_settings and not current_user.role.can_view_all and not current_user.role.can_edit_all:
        raise HTTPException(status_code=403, detail="Finance access required")
    return current_user

def _require_finance_edit(current_user: models.User = Depends(get_current_user)):
    if not getattr(current_user.role, 'can_edit_finance', False) and not current_user.role.can_manage_settings and not current_user.role.can_edit_all:
        raise HTTPException(status_code=403, detail="Finance edit permission required")
    return current_user

def _get_root_service(db: Session, dept_id: int):
    """Climbs up the department tree to find the first ancestor with category=1 (Service)."""
    current_dept = db.query(models.Department).filter(models.Department.id == dept_id).first()
    while current_dept:
        if current_dept.category == 1:
            return current_dept
        if current_dept.parent_id is None:
            break
        current_dept = db.query(models.Department).filter(models.Department.id == current_dept.parent_id).first()
    return None

def _write_audit(db: Session, user: models.User, action: str, target: str,
                 old_value: str = None, new_value: str = None):
    log = models.FinanceAuditLog(
        user_id=user.id, action=action, target=target,
        old_value=old_value, new_value=new_value,
        timestamp=datetime.utcnow()
    )
    db.add(log)


@app.get("/api/salary-rates")
def get_salary_rates(db: Session = Depends(get_db),
                     current_user: models.User = Depends(_require_finance_view)):
    rates = db.query(models.SalaryRate).all()
    return [{"id": r.id, "dept_id": r.dept_id,
             "dept_name": r.department.full_name if r.department else None,
             "position_id": r.position_id,
             "position_name": r.position.name if r.position else None,
             "hourly_rate": r.hourly_rate} for r in rates]


@app.post("/api/salary-rates")
def upsert_salary_rate(rate: SalaryRateCreate, db: Session = Depends(get_db),
                       current_user: models.User = Depends(_require_finance_edit)):
    existing = db.query(models.SalaryRate).filter(
        models.SalaryRate.dept_id == rate.dept_id,
        models.SalaryRate.position_id == rate.position_id
    ).first()
    pos = db.query(models.Position).get(rate.position_id)
    dept = db.query(models.Department).get(rate.dept_id)
    target = f"{pos.name if pos else rate.position_id} in {dept.name if dept else rate.dept_id}"
    if existing:
        old_val = str(existing.hourly_rate)
        existing.hourly_rate = rate.hourly_rate
        _write_audit(db, current_user, "UPDATE_RATE", target, old_val, str(rate.hourly_rate))
        db.commit()
        db.refresh(existing)
        r = existing
    else:
        r = models.SalaryRate(**rate.dict())
        db.add(r)
        _write_audit(db, current_user, "CREATE_RATE", target, None, str(rate.hourly_rate))
        db.commit()
        db.refresh(r)
    return {"id": r.id, "dept_id": r.dept_id, "position_id": r.position_id, "hourly_rate": r.hourly_rate}


@app.delete("/api/salary-rates/{rate_id}")
def delete_salary_rate(rate_id: int, db: Session = Depends(get_db),
                       current_user: models.User = Depends(_require_finance_edit)):
    r = db.query(models.SalaryRate).filter(models.SalaryRate.id == rate_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Rate not found")
    target = f"position_id={r.position_id} dept_id={r.dept_id}"
    _write_audit(db, current_user, "DELETE_RATE", target, str(r.hourly_rate), None)
    db.delete(r)
    db.commit()
    return {"status": "deleted"}


@app.get("/api/finance/payroll/{year_month}")
def get_payroll(year_month: str, db: Session = Depends(get_db),
                current_user: models.User = Depends(_require_finance_view)):
    try:
        year, month = map(int, year_month.split("-"))
    except ValueError:
        raise HTTPException(status_code=400, detail="year_month must be YYYY-MM")
    _, days = calendar.monthrange(year, month)
    month_start = date(year, month, 1)
    month_end = date(year, month, days)
    rates = db.query(models.SalaryRate).all()
    rate_map = {(r.dept_id, r.position_id): r.hourly_rate for r in rates}
    employees = db.query(models.Employee).all()
    rows = []
    dept_totals: Dict[int, dict] = {}
    for emp in employees:
        entries = db.query(models.Timesheet).filter(
            models.Timesheet.employee_id == emp.id,
            models.Timesheet.date >= month_start,
            models.Timesheet.date <= month_end,
        ).all()
        rate = rate_map.get((emp.dept_id, emp.position_id), 0.0)
        total_std = 0.0
        total_night = 0.0
        gross_pay = 0.0
        for entry in entries:
            wc = db.query(models.WorkCode).get(entry.work_code_id)
            if wc:
                m = wc.rate_multiplier or 1.0
                total_std += wc.hours_standard
                total_night += wc.hours_night
                gross_pay += (wc.hours_standard + wc.hours_night) * rate * m
        dept_name = emp.department.name if emp.department else "Unknown"
        service = _get_root_service(db, emp.dept_id)
        service_name = service.name if service else dept_name
        service_id = service.id if service else emp.dept_id

        pos_name = emp.position.name if emp.position else "—"
        pos_category = emp.category if emp.category is not None else 99
        rows.append({
            "employee_id": emp.id, "full_name": emp.full_name,
            "tab_number": emp.tab_number, "position": pos_name,
            "category": pos_category,
            "dept_id": emp.dept_id, "dept_name": dept_name,
            "service_id": service_id, "service_name": service_name,
            "hourly_rate": rate,
            "std_hours": round(total_std, 1), "night_hours": round(total_night, 1),
            "total_hours": round(total_std + total_night, 1),
            "gross_pay": round(gross_pay, 2),
        })
        if emp.dept_id not in dept_totals:
            dept_totals[emp.dept_id] = {"dept_name": dept_name, "total_pay": 0.0, "employees": 0}
        dept_totals[emp.dept_id]["total_pay"] += gross_pay
        dept_totals[emp.dept_id]["employees"] += 1
    rows.sort(key=lambda x: (x["service_name"], x["dept_name"], x["category"]))
    grand_total = sum(r["gross_pay"] for r in rows)
    avg_salary = round(grand_total / len(rows), 2) if rows else 0.0
    dept_summary = [{"dept_id": did, **info, "total_pay": round(info["total_pay"], 2)}
                    for did, info in dept_totals.items()]
    top = max(dept_summary, key=lambda x: x["total_pay"]) if dept_summary else None
    return {
        "year_month": year_month, "employees": rows, "dept_summary": dept_summary,
        "grand_total": round(grand_total, 2), "avg_salary": avg_salary,
        "top_dept": top["dept_name"] if top else None,
        "top_dept_pay": round(top["total_pay"], 2) if top else 0.0,
    }


@app.get("/api/finance/payroll/{year_month}/export")
def export_payroll_excel(year_month: str, db: Session = Depends(get_db),
                         current_user: models.User = Depends(_require_finance_view)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    payroll = get_payroll(year_month, db, current_user)
    wb = Workbook()
    ws = wb.active
    ws.title = f"Payroll {year_month}"
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(fill_type="solid", fgColor="4F46E5")
    sfill = PatternFill(fill_type="solid", fgColor="E2E8F0")
    sfont = Font(bold=True)
    alt = PatternFill(fill_type="solid", fgColor="F1F5F9")
    hdrs = ["#", "Tab No.", "Full Name", "Position", "Department",
            "Rate/hr", "Std Hrs", "Night Hrs", "Total Hrs", "Gross Pay"]
    ws.append(hdrs)
    for c in range(1, len(hdrs) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hf; cell.fill = hfill
        cell.alignment = Alignment(horizontal="center"); cell.border = border
    
    current_service = None
    row_ptr = 2
    for i, emp in enumerate(payroll["employees"], 1):
        if emp["service_name"] != current_service:
            current_service = emp["service_name"]
            ws.append([f"SERVICE: {current_service}"])
            ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=10)
            cell = ws.cell(row=row_ptr, column=1)
            cell.font = sfont; cell.fill = sfill; cell.border = border; cell.alignment = Alignment(horizontal="left")
            row_ptr += 1

        rdata = [i, emp["tab_number"], emp["full_name"], emp["position"], emp["dept_name"],
                 emp["hourly_rate"], emp["std_hours"], emp["night_hours"],
                 emp["total_hours"], emp["gross_pay"]]
        ws.append(rdata)
        fill = alt if i % 2 == 0 else None
        for c in range(1, len(rdata) + 1):
            cell = ws.cell(row=row_ptr, column=c)
            cell.border = border
            if fill: cell.fill = fill
            if c == 10: cell.number_format = '#,##0.00'
        row_ptr += 1

    tr = ws.max_row + 1
    ws.cell(tr, 1, "TOTAL").font = Font(bold=True)
    tc = ws.cell(tr, 10, payroll["grand_total"])
    tc.font = Font(bold=True); tc.number_format = '#,##0.00'
    tc.fill = PatternFill(fill_type="solid", fgColor="C7D2FE")
    for c, w in enumerate([4, 10, 28, 20, 20, 10, 9, 10, 10, 14], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=payroll_{year_month}.xlsx"})


@app.get("/api/finance/audit-log")
def get_audit_log(db: Session = Depends(get_db),
                  current_user: models.User = Depends(_require_finance_edit)):
    logs = db.query(models.FinanceAuditLog).order_by(
        models.FinanceAuditLog.timestamp.desc()).limit(200).all()
    return [{
        "id": log.id,
        "username": log.user.username if log.user else "system",
        "action": log.action, "target": log.target,
        "old_value": log.old_value, "new_value": log.new_value,
        "timestamp": log.timestamp.isoformat() if log.timestamp else None,
    } for log in logs]
