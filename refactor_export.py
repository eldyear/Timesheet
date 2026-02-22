import re

with open('backend/main.py', 'r') as f:
    content = f.read()

# Define the exact start and end pattern for replacement
start_pattern = r'@app\.get\("/api/export/t13/{dept_id}/{year_month}"\)\ndef export_t13\(dept_id: int, year_month: str, db: Session = Depends\(get_db\), current_user: models\.User = Depends\(get_current_user\)\):'
end_pattern = r'    return StreamingResponse\(\n *output,\n *media_type="application/vnd\.openxmlformats-officedocument\.spreadsheetml\.sheet",\n *headers={"Content-Disposition": f"attachment; filename={filename}"}\n *\)'

match = re.search(f"{start_pattern}.*?{end_pattern}", content, re.DOTALL)
if not match:
    print("Could not find the export_t13 function to replace")
    exit(1)

new_func = """@app.get("/api/export/t13/{dept_id}/{year_month}")
def export_t13(dept_id: int, year_month: str, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    \"\"\"Exports Timesheet to Excel T-13 Format with Department Grouping\"\"\"
    if not current_user.role.can_view_all and not current_user.role.can_edit_all and not current_user.role.can_view_only:
        allowed_dept_ids = _get_department_hierarchy_ids(db, current_user.active_dept_id)
        if dept_id not in allowed_dept_ids:
            raise HTTPException(status_code=403, detail="Not authorized to export this department's timesheet")

    try:
        year, month = map(int, year_month.split("-"))
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid month format. Expected YYYY-MM")

    department = db.query(models.Department).filter(models.Department.id == dept_id).first()
    if not department:
        raise HTTPException(status_code=404, detail="Department not found")

    dept_ids = _get_department_hierarchy_ids(db, dept_id)
    departments = db.query(models.Department).filter(models.Department.id.in_(dept_ids)).all()
    employees = db.query(models.Employee).filter(models.Employee.dept_id.in_(dept_ids)).all()
    emp_ids = [emp.id for emp in employees]

    _, last_day = calendar.monthrange(year, month)
    start_date = date(year, month, 1)
    end_date = date(year, month, last_day)

    entries = db.query(models.Timesheet).filter(
        models.Timesheet.employee_id.in_(emp_ids),
        models.Timesheet.date >= start_date,
        models.Timesheet.date <= end_date
    ).all()

    work_codes = {wc.id: wc for wc in db.query(models.WorkCode).all()}
    
    # Timesheet Map: emp.id -> day -> wc
    timesheet_map = {emp.id: {} for emp in employees}
    for entry in entries:
        timesheet_map[entry.employee_id][entry.date.day] = work_codes.get(entry.work_code_id)

    # Calculate employee totals
    totals = {}
    for emp in employees:
        std = 0.0
        night = 0.0
        for day in range(1, last_day + 1):
            wc = timesheet_map[emp.id].get(day)
            if wc:
                std += (wc.hours_standard or 0.0)
                night += (wc.hours_night or 0.0)
        totals[emp.id] = {"std": round(std, 1), "night": round(night, 1), "total": round(std + night, 1)}

    # Group employees by department for rendering
    depts_map = {d.id: d for d in departments}
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from urllib.parse import quote
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Timesheet {year_month}"
    
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Header Font and Fill
    hf = Font(bold=True, color="1E293B")
    hfill = PatternFill(fill_type="solid", fgColor="F8FAFC")
    
    # Headers
    hdrs = ["Employee Name", "Tab No.", "Position", "Category"]
    for day in range(1, last_day + 1):
        hdrs.append(str(day))
    hdrs.extend(["Std Hrs", "Night Hrs", "Total Hrs"])
    
    ws.append(hdrs)
    for c in range(1, len(hdrs) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Process Departments Top-Down
    # Sort departments by category, then by ID
    sorted_depts = sorted(departments, key=lambda d: (d.category or 99, d.id))
    
    current_row = 2
    
    for dept in sorted_depts:
        dept_emps = [e for e in employees if e.dept_id == dept.id]
        if not dept_emps:
            continue
            
        # Draw Department Banner
        dept_cell = ws.cell(row=current_row, column=1)
        
        parent_name = ""
        if dept.parent_id and dept.parent_id in depts_map:
            parent_name = f"{depts_map[dept.parent_id].name} » "
            
        dept_cell.value = f"{parent_name}{dept.name}"
        dept_cell.font = Font(bold=True, color="0F172A")
        dept_cell.fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
        
        # Merge department banner across all columns
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(hdrs))
        # Apply borders to merged cell
        for c in range(1, len(hdrs) + 1):
            ws.cell(row=current_row, column=c).border = border
            
        current_row += 1
        
        # Draw Employees
        dept_emps.sort(key=lambda e: (e.category or 99, e.full_name))
        for emp in dept_emps:
            pos_name = emp.position.name if emp.position else "\u2014"
            cat_name = str(emp.category) if emp.category is not None else "99"
            row_data = [emp.full_name, emp.tab_number, pos_name, cat_name]
            
            for day in range(1, last_day + 1):
                wc = timesheet_map[emp.id].get(day)
                row_data.append(wc.code if wc else "")
                
            emp_totals = totals[emp.id]
            row_data.extend([emp_totals["std"], emp_totals["night"], emp_totals["total"]])
            
            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=c_idx)
                cell.value = val
                cell.border = border
                if c_idx > 4:  # Center align days and totals
                    cell.alignment = Alignment(horizontal="center")
                    
            current_row += 1

    # Adjust Column Widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    for day in range(1, last_day + 1):
        ws.column_dimensions[get_column_letter(day + 4)].width = 5
    for c in range(1, 4):
        ws.column_dimensions[get_column_letter(last_day + 4 + c)].width = 10

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"T-13_{department.full_name.replace(' ', '_').replace('»', '-')}_{year_month}.xlsx"
    encoded_filename = quote(filename)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=utf-8''{encoded_filename}"}
    )"""

new_content = content.replace(match.group(0), new_func)

with open('backend/main.py', 'w') as f:
    f.write(new_content)
print("Updated main.py successfully")
